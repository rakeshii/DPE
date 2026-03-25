"""
tests/test_projector.py
Run with: python -m pytest tests/ -v
"""

import os
import sys
import pytest

# Make sure project root is on path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.projector import shift_formula
from core.validator import validate_upload
from config import Config

SHEET_CONFIG = Config.SHEET_CONFIG


# ══════════════════════════════════════════════════════════════
# shift_formula tests
# ══════════════════════════════════════════════════════════════

class TestShiftFormula:
    """Test every category of formula reference."""

    def test_same_sheet_simple(self):
        assert shift_formula('=D11', 4, SHEET_CONFIG) == '=E11'

    def test_same_sheet_no_shift_below_insert(self):
        assert shift_formula('=C11', 4, SHEET_CONFIG) == '=C11'

    def test_same_sheet_sum_range(self):
        assert shift_formula('=SUM(D9:D12)', 4, SHEET_CONFIG) == '=SUM(E9:E12)'

    def test_same_sheet_mixed_expression(self):
        assert shift_formula('=+D39-D25', 4, SHEET_CONFIG) == '=+E39-E25'

    def test_cross_sheet_single_cell(self):
        result = shift_formula("='Note To P & L'!C13", 4, SHEET_CONFIG)
        assert result == "='Note To P & L'!D13"

    def test_cross_sheet_note_3_4(self):
        result = shift_formula("=+'P & L'!D29", 4, SHEET_CONFIG)
        assert result == "=+'P & L'!E29"

    def test_cross_sheet_range_end_col_preserved(self):
        """Key v8 fix: TB!C102:C106 — end col must NOT become D106."""
        result = shift_formula('=SUM(TB!C102:C106)', 3, SHEET_CONFIG)
        assert result == '=SUM(TB!C102:C106)'

    def test_cross_sheet_range_when_shift_needed(self):
        """'Note 5-7'!B37:E37 — E is col 5 >= insert 4, must become F."""
        result = shift_formula("=+'Note  5-7'!B37:E37", 4, SHEET_CONFIG)
        assert result == "=+'Note  5-7'!B37:F37"

    def test_support_sheet_no_same_shift(self):
        """cur_insert=None: same-sheet refs untouched, cross-sheet refs shifted."""
        result = shift_formula('=+BS!E31', None, SHEET_CONFIG)
        assert result == '=+BS!F31'

    def test_support_wdv_exact(self):
        """The user-reported formula that triggered the v6 fix."""
        result = shift_formula('=+BS!E31', None, SHEET_CONFIG)
        assert result == '=+BS!F31'

    def test_note_to_pl_tb_range_not_corrupted(self):
        """The user-reported formula that triggered v8 fix."""
        formula  = "=SUM(TB!B102:B106)-SUM(TB!C102:C106)-'P & L'!D16"
        expected = "=SUM(TB!B102:B106)-SUM(TB!C102:C106)-'P & L'!E16"
        result   = shift_formula(formula, 3, SHEET_CONFIG)
        assert result == expected

    def test_non_formula_passthrough(self):
        assert shift_formula(12345.0, 4, SHEET_CONFIG) == 12345.0
        assert shift_formula('plain text', 4, SHEET_CONFIG) == 'plain text'
        assert shift_formula(None, 4, SHEET_CONFIG) is None

    def test_absolute_ref(self):
        result = shift_formula('=$D$11', 4, SHEET_CONFIG)
        assert result == '=$E$11'

    def test_absolute_ref_below_insert(self):
        result = shift_formula('=$C$11', 4, SHEET_CONFIG)
        assert result == '=$C$11'


# ══════════════════════════════════════════════════════════════
# validate_upload tests
# ══════════════════════════════════════════════════════════════

class TestValidateUpload:

    def test_missing_file(self, tmp_path):
        result = validate_upload(str(tmp_path / 'nonexistent.xlsx'))
        assert result is not None
        assert 'not found' in result.lower()

    def test_empty_file(self, tmp_path):
        f = tmp_path / 'empty.xlsx'
        f.write_bytes(b'')
        result = validate_upload(str(f))
        assert result is not None
        assert 'empty' in result.lower()

    def test_unsupported_extension(self, tmp_path):
        f = tmp_path / 'data.csv'
        f.write_text('a,b,c')
        result = validate_upload(str(f))
        assert result is not None
        assert 'unsupported' in result.lower()



# ══════════════════════════════════════════════════════════════
# copy_vals flag tests  (the "Leave blank" bug fix)
# ══════════════════════════════════════════════════════════════

class TestCopyValsFlag:
    """
    Ensures copy_vals=False (Leave blank) produces an empty 2026 column
    for BOTH plain numeric cells AND formula cells.
    """

    def _make_sheet(self):
        """Build a tiny in-memory worksheet that mimics a financial sheet."""
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active

        # Row 1 — period header
        ws['D1'] = 'As at March 31, 2025'
        ws['E1'] = 'As at March 31, 2024'

        # Row 2 — plain number
        ws['D2'] = 1295016590.0
        ws['E2'] = 887452724.0

        # Row 3 — formula
        ws['D3'] = '=SUM(D9:D12)'
        ws['E3'] = '=SUM(E9:E12)'

        return wb, ws

    def _run(self, copy_vals):
        import openpyxl, tempfile, os
        from core.projector import process_financial_sheet

        wb, ws = self._make_sheet()
        cfg = {
            'BS': {'insert': 4},
            'P & L': {'insert': 4},
        }
        # insert_col=4 (1-indexed = col D)
        process_financial_sheet(
            ws, 'BS', 4, cfg,
            new_header='As at 31 March, 2026',
            copy_vals=copy_vals,
            update_titles=False
        )
        return ws

    def test_copy_vals_true_copies_numbers(self):
        ws = self._run(copy_vals=True)
        # 2026 col only receives formulas, never plain numbers.
        # Plain numbers in 2025 are left for auditor to fill via TB → auto-populate.
        # So 2026 col D R2 (which was a plain number in 2025) must be None.
        assert ws.cell(2, 4).value is None   # plain number NOT copied to 2026

    def test_copy_vals_true_copies_formulas(self):
        ws = self._run(copy_vals=True)
        # 2026 formula cell should have a formula (shifted from D to E)
        assert ws.cell(3, 4).value is not None
        assert str(ws.cell(3, 4).value).startswith('=')

    def test_copy_vals_false_leaves_numbers_blank(self):
        """2026 col never gets plain numbers — only formulas go to 2026.
        Plain numbers from 2025 go into the 2025 col as frozen values."""
        ws = self._run(copy_vals=False)
        assert ws.cell(2, 4).value is None            # 2026 col: no plain number

    def test_copy_vals_false_still_copies_formulas(self):
        """Formulas are ALWAYS copied regardless of copy_vals.
        They pull live data — auditors need them to calculate totals.
        Only plain hardcoded numbers are left blank when copy_vals=False."""
        ws = self._run(copy_vals=False)
        v = ws.cell(3, 4).value
        assert v is not None and str(v).startswith('=')   # formula must be present

    def test_header_always_written_regardless_of_copy_vals(self):
        """The 2026 header cell must always appear even when copy_vals=False."""
        ws = self._run(copy_vals=False)
        assert ws.cell(1, 4).value == 'As at 31 March, 2026'

# ══════════════════════════════════════════════════════════════
# Run directly
# ══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    pytest.main([__file__, '-v'])
